import pandas as pd
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from web_click import web_click

def set_prompt(court_name):
    return f"""
    搜索一下。搜索一下。使用百度，谷歌等搜索引擎搜索。
    请搜索{court_name}的领导班子组成。
    请以以下格式进行输出。
    如果公开资料个人信息较少也不要不填写，查找不到的部分写NA即可。
    所在法院：xx；领导姓名，：xx；领导在法院的职务：xx；
    领导目前除在法院任职外，在人大、政协等其他政法机构的同步任职情况（如不存在填写无）：xxx；
    领导性别：xx；领导籍贯：xx；领导出生年份：xx；领导受教育经历：xx；领导最高学历：xx;
    领导上任前在其他职位的历史任职信息：xx；
    信息依据的网站地址：xx。
    得到结果后以json格式输出，必须返回严格的JSON格式。不要返回不相关的内容。
    信息依据的网站地址只需要返回地址url即可
    """

def process_excel(file_path):
    # 读取Excel文件，明确指定只读取sheet1
    df = pd.read_excel(file_path, sheet_name="Sheet1")
    
    # 存储结果的列表
    results = []
    
    # 从第3行开始遍历(实际上因为pandas的索引从0开始，所以是索引2开始)
    for i in range(2, len(df)):
        # 获取当前行的信息
        province = df.iloc[i, 1]  # 所在省 (B列)
        city = df.iloc[i, 2]      # 隶属关系上级市 (C列)
        higher_court = df.iloc[i, 3]  # 隶属关系上级法院 (D列)
        court_level = df.iloc[i, 4]   # 法院层级 (E列)
        court_name = df.iloc[i, 5]    # 法院名称 (F列)
        
        # 确保所有值都是字符串类型
        province = str(province) if not pd.isna(province) else "NA"
        city = str(city) if not pd.isna(city) else "NA"
        higher_court = str(higher_court) if not pd.isna(higher_court) else "NA"
        court_level = str(court_level) if not pd.isna(court_level) else "NA"
        court_name = str(court_name) if not pd.isna(court_name) else "NA"
        
        # 跳过没有法院名称的行
        if court_name == "NA" or court_name.strip() == "":
            continue
        
        # 创建包含所有信息的字典
        court_info = {
            "所在省": province,
            "隶属关系上级市": city,
            "隶属关系上级法院": higher_court,
            "法院层级": court_level,
            "法院名称": court_name,
            "prompt": set_prompt(court_name)
        }
        
        results.append(court_info)
    
    return results

def process_and_write_to_excel(file_path):
    # 获取处理后的数据
    courts_data = process_excel(file_path)
    
    try:
        # 尝试打开现有Excel文件
        workbook = openpyxl.load_workbook(file_path)
        # 检查是否已存在Sheet2，如果不存在则创建
        if "Sheet2" not in workbook.sheetnames:
            workbook.create_sheet("Sheet2")
    except FileNotFoundError:
        # 如果文件不存在，创建新的工作簿
        workbook = openpyxl.Workbook()
        workbook.active.title = "Sheet1"
        workbook.create_sheet("Sheet2")
    
    # 获取Sheet2
    sheet2 = workbook["Sheet2"]
    
    # 定义字段顺序对应索引
    FIELD_ORDER = [
        "所在法院",           # 0
        "领导姓名",           # 1
        "领导在法院的职务",    # 2
        "领导目前除在法院任职外，在人大、政协等其他政法机构的同步任职情况", # 3
        "领导性别",           # 4
        "领导籍贯",           # 5
        "领导出生年份",        # 6
        "领导受教育经历",      # 7
        "领导最高学历",        # 8
        "领导上任前在其他职位的历史任职信息", # 9
        "信息依据的网站地址"    # 10
    ]
    
    # 初始行号，从第4行开始
    current_row = 4
    
    # 处理每条court_info记录
    for court_info in courts_data:
        try:
            # 调用web_click函数获取领导信息
            leaders_json = web_click(court_info["prompt"])
            
            # 解析JSON字符串为Python对象（如果返回的是字符串）
            if isinstance(leaders_json, str):
                leaders = json.loads(leaders_json)
            else:
                leaders = leaders_json
            
            # 处理每位领导信息
            for leader in leaders:
                # 将JSON对象转换为按预定义顺序排列的值列表
                values = []
                for field in FIELD_ORDER:
                    values.append(leader.get(field, "NA"))
                
                # 提取领导相关信息
                name = values[1] if len(values) > 1 else "NA"
                position = values[2] if len(values) > 2 else "NA"
                other_roles = values[3] if len(values) > 3 else "无"
                gender = values[4] if len(values) > 4 else "NA"
                birthplace = values[5] if len(values) > 5 else "NA"
                birth_year = values[6] if len(values) > 6 else "NA"
                education = values[7] if len(values) > 7 else "NA"
                highest_education = values[8] if len(values) > 8 else "NA"
                work_history = values[9] if len(values) > 9 else "NA"
                source_url = values[10] if len(values) > 10 else "NA"
                
                # 写入基本法院信息
                sheet2.cell(row=current_row, column=2, value=court_info["法院名称"])
                sheet2.cell(row=current_row, column=3, value=court_info["所在省"])
                sheet2.cell(row=current_row, column=4, value=court_info["隶属关系上级市"])
                sheet2.cell(row=current_row, column=5, value=court_info["隶属关系上级法院"])
                sheet2.cell(row=current_row, column=6, value=court_info["法院层级"])
                
                # 写入领导信息
                sheet2.cell(row=current_row, column=8, value=name)
                sheet2.cell(row=current_row, column=9, value=position)
                sheet2.cell(row=current_row, column=10, value=other_roles)
                sheet2.cell(row=current_row, column=11, value=gender)
                sheet2.cell(row=current_row, column=12, value=birthplace)
                sheet2.cell(row=current_row, column=13, value=birth_year)
                sheet2.cell(row=current_row, column=14, value=education)
                sheet2.cell(row=current_row, column=15, value=highest_education)
                sheet2.cell(row=current_row, column=16, value=work_history)
                sheet2.cell(row=current_row, column=17, value=source_url)
                
                # 增加行号
                current_row += 1
                
        except Exception as e:
            # 记录错误信息
            print(f"处理{court_info['法院名称']}时出错: {str(e)}")
            
            # 仍然写入基本法院信息，但领导信息留空
            sheet2.cell(row=current_row, column=1, value=court_info["法院名称"])
            sheet2.cell(row=current_row, column=2, value=court_info["所在省"])
            sheet2.cell(row=current_row, column=3, value=court_info["隶属关系上级市"])
            sheet2.cell(row=current_row, column=4, value=court_info["隶属关系上级法院"])
            sheet2.cell(row=current_row, column=5, value=court_info["法院层级"])
            sheet2.cell(row=current_row, column=18, value=f"处理出错: {str(e)}")
            
            # 增加行号
            current_row += 1
    
    # 保存Excel文件
    workbook.save(file_path)
    
    print(f"处理完成，数据已写入Sheet2，从第4行开始。")

def main():
    file_path = "data/test0319.xlsx"  # Excel文件路径
    process_and_write_to_excel(file_path)

if __name__ == "__main__":
    main()

